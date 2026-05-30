import os, sys, datetime, time, warnings
os.environ['MPLBACKEND'] = 'Agg'
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import twstock, ta
from tqdm import tqdm
from multiprocessing import Pool, cpu_count, freeze_support
from sklearn.ensemble import RandomForestClassifier 
from sklearn.model_selection import train_test_split 
from sklearn.preprocessing import StandardScaler
from collections import defaultdict
from functools import partial
import multiprocessing as mp
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import requests
import json, openpyxl, csv

# 替換成你自己的 Access Token 和 userId
CHANNEL_ACCESS_TOKEN = '5cq7sPHsT7DZ6IEFilgGhhEkIwiuWep3iItjh7PsIsJaMk9Cxc5VxWkDbTrjl5SLgZHgdWotr2qKcr0WV2pGt7kQ7N4eKOtDjEmO6MkKzichFMwxrNF1sNp2xtgHmI78SoHalXaUyFytAmQOdh1uGgdB04t89/1O/w1cDnyilFU='
USER_ID = 'U2782701b952c83a39fe757d337188d78'

warnings.filterwarnings("ignore")

# 執行時間:
#   銀行業:  56.42,  108.66,  69.16
#   醫療業:  89.06,  123.62,  79.41
#   科技業: 883.57, 1037.76, 878.93
# 傳統產業: 789.45,  813.62, 746.53

# ====== 參數 ======
# industry_dict = {
#     '銀行業':['金融保險業'],
#     '醫療業':['生技醫療業'],
#     '科技業':['電子工業','半導體業','電腦及週邊設備業','光電業','通信網路業','電子零組件業','電子通路業','資訊服務業','其他電子業','電子商務','數位雲端'],     
#     '傳統產業':['航運業','電機機械','水泥工業','電器電纜','鋼鐵工業','汽車工業','食品工業','塑膠工業','紡織纖維','化學工業',
#                '造紙工業','橡膠工業','玻璃陶瓷','建材營造','貿易百貨','油電燃氣業','觀光餐旅','其他','綠能環保'],
#     '其他行業':['文化創意業','農業科技業','綜合','運動休閒','居家生活']            
# }
industry_dict = {
    '科技業':['電子工業','半導體業','電腦及週邊設備業','光電業','通信網路業','電子零組件業','電子通路業','資訊服務業','其他電子業','電子商務','數位雲端'],     
    '傳統產業':['航運業','電機機械','水泥工業','電器電纜','鋼鐵工業','汽車工業','食品工業','塑膠工業','紡織纖維','化學工業',
               '造紙工業','橡膠工業','玻璃陶瓷','建材營造','貿易百貨','油電燃氣業','觀光餐旅','綠能環保'],
}
user_choice = 'all' # 選擇行業別
years = 2 # 要訓練和回測幾年內的股票
lookahead_days = 10 # 幾天內要看到勝敗
profit_target = 0.05 #上漲幾 % 算勝利條件
show_days = 60 # 畫圖要往前畫幾天
vol_times = 6 # 成交量幾倍作為進場訊號
is_plot = False

# ====== Logger ======
class Logger(object):
    def __init__(self, filename):
        self.terminal = sys.stdout
        self.log = open(filename, "w", encoding="utf-8")
    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
    def flush(self):
        self.terminal.flush()
        self.log.flush()

def send_line_message(text):
    url = 'https://api.line.me/v2/bot/message/push'
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {CHANNEL_ACCESS_TOKEN}'
    }
    data = {
        'to': USER_ID,
        'messages': [
            {
                'type': 'text',
                'text': text
            }
        ]
    }
    response = requests.post(url, headers=headers, data=json.dumps(data))
    if response.status_code == 200:
        print("✅ 成功推播訊息到 LINE")
    else:
        print(f"❌ 推播失敗：{response.status_code}, {response.text}")
    return response.text

def setup_logger(log_path):
    """每次都設置 logger，覆蓋 sys.stdout，避免多次疊加"""
    sys.stdout = Logger(log_path)
    print(f"所有 print 內容將同步寫入 {log_path}")

def add_features(df):
    """計算技術指標"""
    df = df.copy()
    # 計算移動平均
    for period, name in zip([5, 20, 60], ['5MA', '20MA', '60MA']):
        df[name] = df['close'].rolling(window=period).mean()
    # 計算技術指標
    df['RSI'] = ta.momentum.RSIIndicator(df['close'], 14).rsi()
    # 計算 MACD
    macd = ta.trend.MACD(df['close'])
    df['MACD'] = macd.macd()
    df['Signal'] = macd.macd_signal()
    df['Histogram'] = macd.macd_diff()
    # 計算 KD
    stoch = ta.momentum.StochasticOscillator(df['high'], df['low'], df['close'], 9, 3)
    df['K'] = stoch.stoch()
    df['D'] = stoch.stoch_signal()
    return df.dropna()

# ====== 型態偵測 ======
def detect_w_bottom(df, i, window=20, tolerance=0.03):
    # 取最近 window 天
    if i < window: return False
    closes = df['close'].iloc[i-window:i+1].values
    lows = df['low'].iloc[i-window:i+1].values 
    # 找兩個低點
    min1_idx = np.argmin(lows[:window//2])
    min2_idx = np.argmin(lows[window//2:]) + window//2
    min1 = lows[min1_idx]
    min2 = lows[min2_idx]
    # 兩個低點接近且中間有反彈
    if abs(min1 - min2) / min1 < tolerance and np.max(closes[min1_idx:min2_idx]) > min1 * (1 + tolerance):
        # 頸線突破
        neckline = np.max(closes[min1_idx:min2_idx])
        if closes[-1] > neckline:
            return True
    return False

def detect_triangle(df, i, window=20, tolerance=0.03):
    """ 偵測最近 window 天是否有三角收斂型態 (高點下降、低點上升) """
    if i < window: return False
    highs = df['high'].iloc[i-window:i+1].values
    lows = df['low'].iloc[i-window:i+1].values
    # 找高點序列和低點序列
    from scipy.signal import argrelextrema
    high_idx = argrelextrema(highs, np.greater, order=2)[0]
    low_idx = argrelextrema(lows, np.less, order=2)[0]

    if len(high_idx) < 2 or len(low_idx) < 2: return False
    # 高點是否下降，低點是否上升
    high_trend = highs[high_idx[-1]] < highs[high_idx[0]]
    low_trend = lows[low_idx[-1]] > lows[low_idx[0]]
    # 收斂條件
    if high_trend and low_trend:
        # 收斂幅度夠大
        if (highs[high_idx[0]] - highs[high_idx[-1]])/highs[high_idx[0]] > tolerance and \
           (lows[low_idx[-1]] - lows[low_idx[0]])/lows[low_idx[0]] > tolerance:
            return True
    return False

def detect_cup_handle(df, i, window=40, cup_depth=0.10, handle_depth=0.05):
    """ 偵測最近 window 天是否有杯柄型態 """
    if i < window: return False
    closes = df['close'].iloc[i-window:i+1].values
    # 杯底
    min_idx = np.argmin(closes)
    left_max = np.max(closes[:min_idx]) if min_idx > 0 else closes[0]
    right_max = np.max(closes[min_idx+1:]) if min_idx < len(closes)-1 else closes[-1]
    cup_height = min(left_max, right_max)
    cup_depth_val = (cup_height - closes[min_idx]) / cup_height
    # 杯底夠深
    if cup_depth_val < cup_depth: return False
    # 杯底後有小回檔 (柄)
    handle_start = min_idx + int((len(closes)-min_idx)*0.5)
    if handle_start >= len(closes)-2: return False
    handle_min = np.min(closes[handle_start:])
    handle_depth_val = (cup_height - handle_min) / cup_height

    if handle_depth_val < handle_depth: return False
    # 柄之後有突破
    if closes[-1] > cup_height: return True
    return False
    
def detect_inverse_head_shoulder(df, i, window=30, tolerance=0.03):
    """ 偵測最近 window 天是否有頭肩底型態 """
    if i < window: return False
    lows = df['low'].iloc[i-window:i+1].values
    # 找三個低點
    from scipy.signal import argrelextrema
    low_idx = argrelextrema(lows, np.less, order=2)[0]
    if len(low_idx) < 3: return False
    # 取最近三個低點
    idxs = low_idx[-3:]
    vals = lows[idxs]
    # 頭肩底條件: 中間最低，兩側接近
    if vals[1] < vals[0] and vals[1] < vals[2]:
        if abs(vals[0] - vals[2]) / vals[1] < tolerance:
            # 頸線突破
            neckline = max(lows[idxs[0]:idxs[2]+1])
            if lows[-1] > neckline:
                return True
    return False

# ====== 型態 function dict (自動化) ======
pattern_funcs = {
    'is_w_buttom': detect_w_bottom,
    'is_triangle': detect_triangle,
    'is_cup_handle': detect_cup_handle,
    'is_inverse_head_shoulder': detect_inverse_head_shoulder
}

# ====== 特徵組合 ======
# 程式會把這些特徵 (如 RSI, MACD, KD, 均線, 成交量等) 組合起來，丟給隨機森林模型訓練。訓練後，模型會根據這些特徵組合，預測"未來幾天內是否會達獲利目標"
def make_features(df, i, n_days=3): 
    try:
        # 先檢查所需欄位是否存在且非 NaN
        # required_cols = ['open','high','low','close','capacity','5MA','20MA','60MA','RSI','MACD','Signal','K','D']
        current, prev = df.iloc[i], df.iloc[i-1]
        # for c in required_cols:
        #     if c not in df.columns:
        #         return None
        #     if pd.isna(current.get(c)) or pd.isna(prev.get(c)):
        #         return None
        
        k_pattern = int(current['close'] > current['open'] and 
                        (current['high'] - current['close']) < (current['close'] - current['open']) and 
                        (current['open'] - current['low']) > 2 * (current['high'] - current['close']))
        ma_cross = int(prev['5MA'] < prev['20MA'] and current['5MA'] > current['20MA'])
        rsi_signal = int(prev['RSI'] < 30 and current['RSI'] > 30)
        macd_signal = int(prev['MACD'] < prev['Signal'] and current['MACD'] > current['Signal'])
        kd_cross = int(prev['K'] < prev['D'] and current['K'] > current['D'] and current['K'] < 30 and current['D'] < 30)
        vol_cross = int(prev['capacity'] > 0 and current['capacity'] >= vol_times * prev['capacity'])
        # 多頭排列
        bullish_ma = int((current['5MA'] > current['20MA']) and (current['20MA'] > current['60MA']))
        prev_bullish_ma = int((prev['5MA'] > prev['20MA']) and (prev['20MA'] > prev['60MA']))
        bullish_ma_just = int(bullish_ma == 1 and prev_bullish_ma == 0)
        vol_up = int(current['capacity'] > df['capacity'].iloc[i-5:i].mean())
        prev_vol_up = int(prev['capacity'] > df['capacity'].iloc[i-6:i-1].mean())
        vol_up_just = int(vol_up == 1 and prev_vol_up == 0)

        rsi_just = int(current['RSI'] > 50 and prev['RSI']  <= 50)
        # 均線上彎
        ma5_up = int(bool(current['5MA'] > prev['5MA']))
        ma20_up = int(bool(current['20MA'] > prev['20MA']))
        ma60_up = int(bool(current['60MA'] > prev['60MA']))
        # 計算技術指標比率
        close_ratios = [current['close'] / (current[ma] + 1e-6) for ma in ['5MA', '20MA', '60MA']]
        vol_ratios = [current['capacity'] / (df['capacity'].iloc[i-w:i].mean() + 1e-6) for w in [5, 20]]
        # 多頭排列/均線上灣持續性
        bullish_ma_n = int(all((df.iloc[j]['5MA'] > df.iloc[j]['20MA'] and df.iloc[j]['20MA'] > df.iloc[j]['60MA']) for j in range(i-n_days+1, i+1))) if i >= n_days else 0
        ma5_up_n = int(all(df.iloc[j]['5MA'] > df.iloc[j-1]['5MA'] for j in range(i-n_days+1, i+1))) if i >= n_days else 0
        # 量能判斷
        vol_up_n = int(all(df.iloc[j]['capacity'] > df['capacity'].iloc[j-n_days:j].mean() for j in range(i-n_days+1, i+1))) if i >= n_days else 0
        # 型態 + 指標複合特徵
        bullish_ma_and_rsi = int(bullish_ma == 1 and current['RSI'] > 50)
        ma5_up_and_macd = int(ma5_up == 1 and macd_signal == 1)
        bullish_ma_and_vol = int(bullish_ma == 1 and vol_up_n == 1)

        pattern_fearures = [int(func(df, i)) for func in pattern_funcs.values()]
        return [k_pattern, ma_cross, rsi_signal, macd_signal, kd_cross, vol_cross,
                current['RSI'], current['MACD'], current['K'], current['D'],
                *close_ratios, *vol_ratios,
                bullish_ma, bullish_ma_just, vol_up_just, rsi_just, ma5_up, ma20_up, ma60_up,
                bullish_ma_n, ma5_up_n, vol_up_n,
                bullish_ma_and_rsi, ma5_up_and_macd, bullish_ma_and_vol,
                *pattern_fearures]
    except Exception as e:
        print(f"特徵計算錯誤: {e}")
        return None

# ====== 特徵名稱自動化 ======
feature_names = [
    'k_pattern','ma_cross','rsi_signal','macd_signal','kd_cross','vol_cross',
    'RSI','MACD','K','D','close/5MA','close/20MA','close/60MA','vol/5MA','vol/20MA',
    'bullish_ma','bullish_ma_just','vol_up_just','rsi_just','ma5_up','ma20_up','ma60_up',
    'bullish_ma_n','ma5_up_n','vol_up_n',
    'bullish_ma_and_rsi','ma5_up_and_macd','bullish_ma_and_vol'
] + list(pattern_funcs.keys())

# ====== 其他 function ======
def make_rf_dataset(df):
    """ 建立訓練資料集 """
    X, y = [], []
    # 計算未來價格變動
    close_prices = df['close'].values
    for i in range(60, len(df)-lookahead_days):
        features = make_features(df, i)
        if features is None or any(pd.isna(f) or np.isinf(f) for f in features):
            continue

        entry_price = close_prices[i] # entry_price 使用開盤價，以符合「開盤後觸發訊號才進場」邏輯
        future_closes = close_prices[i+1:i+1+lookahead_days]
        if len(future_closes) == 0:
            continue

        future_max = np.max(future_closes)
        future_min = np.min(future_closes)
        if future_min <= entry_price * (1 - profit_target):
            win = 0 # 10 天內有跌 5%，直接算敗
        elif future_max >= entry_price * (1 + profit_target):
            win = 1 # 10 天內有漲 5%，且沒先跌 5%，算勝
        else:
            win = 0 # 10 天內都沒漲跌也沒跌 5%，算敗
        X.append(features)
        y.append(win)

    return X, y

def get_filtered_stock(data):
    """ 獲取符合條件的股票列表 """
    return [
        code for code in data
        if hasattr(twstock.codes[code], 'group') and twstock.codes[code].group in industry_dict[user_choice]
        and hasattr(twstock.codes[code], 'market') and twstock.codes[code].market in ['上市','上櫃']
        and not code.endswith('U') and not code.endswith('B') and not code.startswith('7') and not code.startswith('9')
        and not getattr(twstock.codes[code], 'removed_date', None)
    ]

def process_backtest(code, all_df, clf, scaler):
    """ 多進程處理回測 """
    try:
        df = all_df[code]
        results = []
        # 計算價格變動
        close_prices, dates = df['close'].values, df.index
        for i in range(60, len(df) - lookahead_days):
            features = make_features(df, i)
            if features is None or any(pd.isna(f) or np.isinf(f) for f in features):
                continue
            score = clf.predict_proba(scaler.transform([features]))[0][1]
            entry_price, entry_date = close_prices[i], dates[i]
            end_date = entry_date + pd.Timedelta(days=lookahead_days)
            future_mask = (dates > entry_date) & (dates <= end_date)
            future_closes = df.loc[future_mask, 'close']
            future_ma5 = df.loc[future_mask, '5MA']
            if len(future_closes) > 0 and len(future_ma5) == len(future_closes):
                profit_pct_arr = (future_closes - entry_price) / entry_price * 100
                profit_pct = profit_pct_arr.max()
                max_profit_date = profit_pct_arr.idxmax()
                min_profit_pct = profit_pct_arr.min()
                if min_profit_pct <= -profit_target * 100:
                    win = 0 # 10 天內有跌 5%，直接算敗
                elif profit_pct >= profit_target * 100:
                    win = 1 # 10 天內有漲 5%，且沒先跌 5%，算勝
                else:
                    win = 0 # 10 天內都沒漲跌也沒跌 5%，算敗
            else:
                profit_pct, max_profit_date, win = 0, None, 0
            results.append((code, entry_date, score, win, profit_pct, max_profit_date, features.copy()))
        return results
    except Exception as e:
        print(f"處理股票 {code} 的回測時發生錯誤: {e}")
        return []

def process_stock(code, data=None):
    """ 多進程處理單一股票資料 """
    try:
        if data is None:
            return code, None
        df = data[code]
        if df is not None and len(df) > 0:
            return code, add_features(df)
        return code, None
    except Exception as e:
        print(f"處理股票 {code} 時發生錯誤: {e}")
        return code, None

def process_training_data(code, all_df):
    """ 多進程處理訓練資料 """
    try:
        if code in all_df:
            return make_rf_dataset(all_df[code])
        return [], []
    except Exception as e:
        print(f"處理股票 {code} 的訓練資料時發生錯誤: {e}")
        return [], []
    
def safe_scatter(ax, dates, data, col, **kwargs):
    """ 安全的繪製散點圖，處理缺失值 """
    try:
        valid_dates = [d for d in dates if d in data.index]
        y = [float(data.loc[d, col]) for d in valid_dates]
        if valid_dates and len(y) == len(valid_dates):
            ax.scatter(valid_dates, y, **kwargs)
    except Exception as e:
        print(f"繪製點散圖時發生錯誤: {e}")

def get_signal_types(score_95, vol_times):
    """ 獲取訊號類型定義 """
    return [
        ('hammer_dates', 'close', 'o', 'red', 'Hammer Signal'),
        ('ma_dates', 'close', '^', 'blue', 'MA Golden Cross'),
        ('rsi_dates', 'close', 's', 'purple', 'RSI Signal'),
        ('macd_dates', 'close', '*', 'green', 'MACD Signal'),
        ('kd_dates', 'close', 'D', 'gold', 'KD Golden Cross'),
        ('vol_dates', 'close', 'x', 'brown', f'Volume x{vol_times} Signal'),
        ('signal_dates', 'close', 'o', 'none', f'Score >= {score_95:.2f}'),
        ('is_w_bottom_dates', 'close', 'P', 'magenta', 'W bottom'),
        ('is_triangle_dates', 'close', 'v', 'cyan', 'Triangle'),
        ('is_cup_handle_dates', 'close', 'h', 'lime', 'Cup'),
        ('is_inverse_head_shoulder_dates', 'close', '8', ' orange', 'Head shoulder'),
    ]

def get_display_width(s):
    s = str(s)
    return sum(2 if ord(c) > 127 else 1 for c in s)

def is_valid_xlsx(path):
    try:
        # 嘗試打開並關閉工作簿
        wb = load_workbook(path)
        wb.close()  # 確保工作簿關閉
        return True
    except Exception as e:
        # 捕捉所有其他錯誤並打印出來
        print(f"檢查檔案 {path} 時發生錯誤：{e}")
        return False

def is_valid_csv(path):
    try:
        # 嘗試讀取 CSV 檔案
        df = pd.read_csv(path)
        return True
    except Exception as e:
        print(f"檢查檔案 {path} 時發生錯誤：{e}")
        return False

def clean_to_float(value):
    try:
        return float(value)
    except ValueError:
        return None  # 如果無法轉換，返回 None

def ensure_column_order(df, desired_order):
    for col in desired_order:
        if col not in df.columns:
            df[col] = ''
    return df[desired_order]

def set_excel_column_width(worksheet, trades):
    for i, col in enumerate(trades.columns):
        max_len = max(trades[col].astype(str).map(get_display_width).max(), get_display_width(col)) + 2
        worksheet.column_dimensions[chr(65+i)].width = max_len

def set_excel_row_fill(ws, state_col_idx, fill_dict, default_fill):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        state = str(row[state_col_idx - 1].value).strip()
        fill = fill_dict.get(state, default_fill)
        for cell in row:
            cell.fill = fill

def next_trading_day(date):
    # date 必須是 datetime 或 Timestamp
    next_day = pd.to_datetime(date) + pd.Timedelta(days=1)
    while next_day.weekday() >= 5: # 5=Saturday, 6=Sunday
        next_day += pd.Timedelta(days=1)
    return next_day

def update_trade_excel(today_signal_list, all_df, excel_path, sheet_name, profit_target=0.05, lookahead_days=10, all_signals_10d=None, all_signals_today=None):

    # 讀取現有 Excel 的該 sheet (若不存在則新建)
    if os.path.exists(excel_path) and not is_valid_xlsx(excel_path):
        print(f"檢查到 {excel_path} 不是合法 xlsx，將自動重建！")
        os.remove(excel_path)

    try:
        if os.path.exists(excel_path):
            trades = pd.read_excel(excel_path, sheet_name=sheet_name)
        else:
            trades = pd.DataFrame(columns=[
                'stock_id', 'stock_name', 'Entry date', 'Entry price', 'Exit date', 'Exit price',
                'Lowest Price', 'Lowest Price in 10 days', 'Highest Price in 10 days',
                'Risk-reward ratio', 'Exit Profit/Loss (%)', 'state'
            ])
    except Exception as e:
        print(f"Error loading file {excel_path}: {e}")
        trades = pd.DataFrame(columns=[
            'stock_id', 'stock_name', 'Entry date', 'Entry price', 'Exit date', 'Exit price',
            'Lowest Price', 'Lowest Price in 10 days', 'Highest Price in 10 days',
            'Risk-reward ratio', 'Exit Profit/Loss (%)', 'state'
        ])
    
    # 強制將數字欄位轉換為 float，並清除非數字數據
    def clean_to_float(value):
        try:
            return float(value)
        except ValueError:
            return np.nan  # 如果無法轉換，返回 NaN
        
    # 檢查所有數值欄位，並強制轉換為數字，清除非數字字符
    trades['Exit Profit/Loss (%)'] = trades['Exit Profit/Loss (%)'].apply(clean_to_float)
    trades['Entry price'] = trades['Entry price'].apply(clean_to_float)
    trades['Exit price'] = trades['Exit price'].apply(clean_to_float)

    # 填補 NaN 為 0 或其他合適的值
    trades.fillna({
        'Exit Profit/Loss (%)': 0.0,
        'Entry price': 0.0,
        'Exit price': 0.0,
        'Lowest Price': 0.0,
        'Lowest Price in 10 days': 0.0,
        'Highest Price in 10 days': 0.0,
        'Risk-reward ratio': 0.0
    }, inplace=True)

    # 補充必要的欄位並處理其他欄位
    trades['stock_id'] = trades['stock_id'].astype(str)
    trades['state'] = trades['state'].str.strip()

    # 先更新部位出場狀態
    for idx, row in trades.iterrows():
        stock_id = str(row['stock_id'])
        df = all_df.get(stock_id)
        if df is None:
            continue
        entry_date = pd.to_datetime(row['Entry date'])
        entry_price = row['Entry price']
        mask = (df.index >= entry_date) & (df.index <= entry_date + pd.Timedelta(days=lookahead_days))
        period_df = df.loc[mask]
        if period_df.empty:
            continue
        min_price_10d = period_df['low'].min() if 'low' in period_df else period_df['close'].min()
        max_price_10d = period_df['high'].max() if 'high' in period_df else period_df['close'].max()
        trades.loc[idx, 'Lowest Price in 10 days'] = min_price_10d
        trades.loc[idx, 'Highest Price in 10 days'] = max_price_10d
        mask_min = (df.index > entry_date)
        min_price = min(row.get('Lowest Price', entry_price), df.loc[mask_min, 'low'].min()) if not pd.isna(row.get('Lowest Price')) else df.loc[mask_min, 'low'].min()
        trades.loc[idx, 'Lowest Price'] = min_price
        
        # 出場判斷
        if row['state'] == 'Holding':
            # 檢查是否已達獲利目標或已過追蹤天數
            if max_price_10d >= entry_price * (1 + profit_target):
                exit_row = period_df[period_df['high'] >= entry_price * (1 + profit_target)].iloc[0]
                trades.loc[idx, ['Exit date', 'Exit price', 'Exit Profit/Loss (%)', 'state']] = [exit_row.name.date(), exit_row['high'], (exit_row['high'] - entry_price) / entry_price * 100, 'Appeared']
            elif (df.index[-1] > entry_date + pd.Timedelta(days=lookahead_days)):  # 若已到期但沒達標，也可以自訂強制出場
                last_row = period_df.iloc[-1]
                trades.loc[idx, ['Exit date', 'Exit price', 'Exit Profit/Loss (%)', 'state']] = [last_row.name.date(), exit_row['close'], (exit_row['close'] - entry_price) / entry_price * 100, 'Appeared']

    # Predicted 轉 Holding，並即時更新所有 item
    all_max_dates = [df.index.max().date() for df in all_df.values() if len(df) > 0]
    today= max(all_max_dates) if all_max_dates else None
    
    if today:
        for idx, row in trades.iterrows():
            if row['state'] == 'Predicted' and row['Entry date'] <= today:
                stock_id = str(row['stock_id'])
                df = all_df.get(stock_id)
                if df is not None and row['Entry date'] in df.index.date:
                    temp_trades = trades.copy()
                    temp_trades.loc[idx, 'state'] = 'Holding'
                    has_unclosed_holding = ((temp_trades['stock_id'].astype(str) == stock_id) & (temp_trades['state'].str.strip() == 'Holding')).any()
                    if has_unclosed_holding:
                        continue
                    entry_price = df.loc[df.index[df.index.date == row['Entry date']], 'open'].values[0]
                    mask = (df.index >= pd.to_datetime(row['Entry date'])) & (df.index <= pd.to_datetime(row['Entry date']) + pd.Timedelta(days=lookahead_days))
                    period_df = df.loc[mask]
                    min_price_10d = period_df['low'].min() if not period_df.empty else entry_price
                    max_price_10d = period_df['high'].max() if not period_df.empty else entry_price
                    trades.loc[idx, ['Entry price', 'Lowest Price', 'Lowest Price in 10 days', 'Highest Price in 10 days']] = [entry_price, entry_price, min_price_10d, max_price_10d]
                    trades.loc[idx, ['Risk-reward ratio', 'Exit Profit/Loss (%)']] = [None, None]
                    trades.loc[idx, 'state'] = 'Holding'
                # else: 不要轉'持有中'，保持 Predicted 狀態

    # 讀取舊 Excel，則補上 stock name 欄位
    if 'stock_name' not in trades.columns:
        trades['stock_name'] = trades['stock_id'].apply(
            lambda x: twstock.codes.get(str(x), None).name if str(x) in twstock.codes else ''
        )

    # 統一計算風報比
    for idx, row in trades.iterrows():
        if row['state'] == 'Holding':
            entry_price = row['Entry price']
            exit_price = row['Exit price']
            min_price_10d = row['Lowest Price in 10 days']
            if pd.notna(exit_price) and pd.notna(min_price_10d) and (entry_price - min_price_10d) != 0:
                trades.loc[idx, 'Risk-reward ratio'] = (exit_price - entry_price) / (entry_price - min_price_10d)
            else:
                trades.loc[idx, 'Risk-reward ratio'] = 0
        else:
            trades.loc[idx, 'Risk-reward ratio'] = None
    
    # 新增今日新訊號 (用今日開盤價進場)
    new_rows = []
    for stock_id, entry_date, score, entry_price in today_signal_list:
        stock_id_str = str(stock_id)
        # 先組合 trades + new_rows (已經準備要新增的部位也要一起判斷)
        temp_trades = pd.concat([trades, pd.DataFrame(new_rows)], ignore_index=True) if new_rows else trades
        has_unclosed_holding = ((temp_trades['stock_id'].astype(str) == stock_id_str) & (temp_trades['state'].str.strip() == 'Holding')).any()
        if has_unclosed_holding:
            continue
        df = all_df[stock_id]
        mask = (df.index >= entry_date) & (df.index <= entry_date + pd.Timedelta(days=lookahead_days))
        period_df = df.loc[mask]
        min_price_10d = period_df['low'].min() if not period_df.empty else entry_price
        max_price_10d = period_df['high'].max() if not period_df.empty else entry_price
        stock_name = twstock.codes.get(str(stock_id_str), None).name if str(stock_id_str) in twstock.codes else ''
        new_rows.append({
            'stock_id': stock_id,
            'stock_name': stock_name,
            'Entry date': entry_date.date(),
            'Entry price': entry_price,
            'Exit date': None,
            'Exit price': None,
            'Lowest Price': entry_price, # 新進場時最低價就是入場價
            'Lowest Price in 10 days': min_price_10d,
            'Highest Price in 10 days': max_price_10d,
            'Risk-reward ratio': None,
            'Exit Profit/Loss (%)': None,
            'state': 'Holding'
        })
    if new_rows:
        trades = pd.concat([trades, pd.DataFrame(new_rows)], ignore_index=True)
        trades.drop_duplicates(subset=['stock_id', 'Entry date', 'state'], keep='first', inplace=True) # 防呆:去除重複
    
    # 存檔前去除時間
    trades['Entry date'] = pd.to_datetime(trades['Entry date']).dt.date
    trades.drop_duplicates(subset=['stock_id', 'Entry date', 'state'], keep='first', inplace=True)

    desired_order = [
        'stock_id','stock_name','Entry date','Entry price','Exit date','Exit price',
        'Lowest Price','Lowest Price in 10 days','Highest Price in 10 days',
        'Risk-reward ratio','Exit Profit/Loss (%)','state'
    ]
    trades = ensure_column_order(trades, desired_order)

    # Excel 寫入
    # with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a' if os.path.exists(excel_path) else 'w', if_sheet_exists='replace') as writer:
    #     trades.to_excel(writer, index=False, sheet_name=sheet_name)

    # 確保 excel_path 存在，並且模式選擇為 append 或 write
    # if os.path.exists(excel_path):
    #     try:
    #         # 嘗試加載現有的 Excel 文件
    #         book = load_workbook(excel_path)
    #     except Exception as e:
    #         print(f"Error loading file {excel_path}: {e}")
    #         # 如果檔案損壞或無法讀取，則創建一個新的 Excel 文件
    #         book = Workbook()
    # else:
    #     # 如果文件不存在，創建新的 Excel 文件
    #     book = Workbook()

    #  # 檢查是否有同名的工作表，並刪除它
    # if sheet_name in book.sheetnames:
    #      del book[sheet_name]

    # # 使用 openpyxl 的 writer，讓 pandas 寫入新的文件
    # with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a' if os.path.exists(excel_path) else 'w') as writer:
    #     # 將打開的工作簿賦值給 writer
    #     writer.book = book
    #     trades.to_excel(writer, index=False, sheet_name=sheet_name)

    # # 用 openpyxl 設定欄寬、底色
    # wb = load_workbook(excel_path)
    # ws = wb[sheet_name]

    # # 設定欄寬
    # for i, col in enumerate(trades.columns):
    #    max_len = max(trades[col].astype(str).map(get_display_width).max(), get_display_width(col)) + 2
    #    ws.column_dimensions[chr(65+i)].width = max_len
        
    # # 底色設定
    # gray_fill = PatternFill(fill_type='solid', fgColor='DDDDDD')
    # blue_full = PatternFill(fill_type='solid', fgColor='CCFFFF')
    # red_fill = PatternFill(fill_type='solid', fgColor='FF9999')
    # no_fill = PatternFill(fill_type=None)
    # header = [cell.value for cell in ws[1]]
    # state_col_idx = header.index('state') + 1
    # fill_dict = {'Appeared': gray_fill,'Predicted': blue_full}
    # set_excel_row_fill(ws, state_col_idx, fill_dict, no_fill)

    # # 額外標記已出場且獲利小於 5% 為紅色
    # profit_col_idx = header.index('Exit Profit/Loss (%)') + 1
    # for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    #    state = str(row[state_col_idx - 1].value).strip()
    #    profit_cell = row[profit_col_idx - 1]
    #    try:
    #        profit = float(profit_cell.value)
    #    except (TypeError, ValueError):
    #        profit = None
    #    if state == 'Appeared' and profit is not None and profit < 5:
    #        for cell in row:
    #            cell.fill = red_fill

    # wb.save(excel_path)
    # wb.close()
    print(f"\n已更新模擬自動下單 Excel: {excel_path}，sheet: {sheet_name}")

def plot_signals(df, clf, scaler, score_95, save_path, show_days, vol_times):
    signal_types = get_signal_types(score_95, vol_times)
    signals = {name: [] for name, *_ in signal_types}
    for i in range(60, len(df)):
        f = make_features(df, i)
        if f is None: continue
        s = clf.predict_proba(scaler.transform([f]))[0][1]
        k_pattern, ma_cross, rsi_signal, macd_signal, kd_cross, vol_cross = f[:6]
        pattern_values= f[-len(pattern_funcs):]
        if s >= score_95: signals['signal_dates'].append(df.index[i])
        if k_pattern: signals['hammer_dates'].append(df.index[i])
        if ma_cross: signals['ma_dates'].append(df.index[i])
        if rsi_signal: signals['rsi_dates'].append(df.index[i])
        if macd_signal: signals['macd_dates'].append(df.index[i])
        if kd_cross: signals['kd_dates'].append(df.index[i])
        if vol_cross: signals['vol_dates'].append(df.index[i])
        for idx, pattern in enumerate(pattern_funcs.keys()):
            if pattern_values[idx]:
                signals[pattern + '_dates'].append(df.index[i])
    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(14, 15), sharex=True, gridspec_kw={'height_ratios':[3, 2, 1]})
    ax1.plot(df.index, df['close'], label='Close Price', color='black')
    ax1.plot(df.index, df['5MA'], label='5MA', color='blue', alpha=0.5)
    ax1.plot(df.index, df['20MA'], label='20MA', color='orange', alpha=0.5)
    for name, col, marker, color, label in signal_types:
        if color == 'none':
            safe_scatter(ax1, signals[name], df, col, marker=marker, s=200, facecolor='none', edgecolor='red', linewidths=2, label=label, zorder=10)
        else:
            safe_scatter(ax1, signals[name], df, col, marker=marker, color=color, label=label, zorder=5)
    ax1.set_title(f"Price Chart with Entry Signals and Weighted Score")
    ax1.set_ylabel('Price')
    ax1.legend(); ax1.grid()
    ax2.plot(df.index, df ['MACD'], label='MACD', color='blue')
    ax2.plot(df.index, df['Signal'], label='Signal', color='orange')
    ax2.bar(df.index, df['Histogram'], label='Histogram', color='gray', alpha=0.3)
    ax2.set_title(f"MACD Indicator")
    ax2.set_ylabel('Value')
    ax2.legend(); ax2.grid()
    ax3.bar(df.index, df['capacity'], color='gray', label='Volume')
    safe_scatter(ax3, signals['vol_dates'], df, 'capacity', marker='x', color='brown', label=f'Volume x{vol_times} Signal', zorder=5)
    safe_scatter(ax3, signals['signal_dates'], df, 'capacity', marker='o', s=200, facecolor='none', edgecolor='red', linewidths=2, label=f'Score >= {score_95:.2f}', zorder=10)
    ax3.set_title(f"Daily Volume")
    ax3.set_xlabel ('Date')
    ax3.set_ylabel('Volume')
    ax3.legend(); ax3.grid()
    plt.tight_layout()
    plt.savefig(save_path)
    plt.close(fig)

# ====== main function ======
def main(save_dir, cache_dir, log_path, global_excel_path, industry):
    start_time = time.time()
    setup_logger(log_path)

    # ====== 讀取大快取 ======
    big_pkl_path = os.path.join(cache_dir, f'all_stock_{years}y.pkl')
    print(f"正在讀取大 pkl: {big_pkl_path}")
    all_data = pd.read_pickle(big_pkl_path)
    print(f"大 pkl 讀取完成，共有 {len(all_data)} 檔股票")

    tw_list = get_filtered_stock(all_data)
    print(f"可用股票數量: {len(tw_list)}")
    print(f"{user_choice} 全部有快取的股票代號: \n{tw_list}")

    # ====== 一次性全部資料與特徵讀入 ======
    process_stock_with_data = partial(process_stock, data=all_data)
    with mp.get_context('spawn').Pool(processes=min(cpu_count(), 4)) as pool:
        results = list(tqdm(pool.imap(process_stock_with_data, tw_list), total=len(tw_list), desc="特徵計算"))
    all_df = {code: df for code, df in results if df is not None}
    if not all_df:
        print("⚠️ 沒有任何股票資料，請檢查快取或條件")
        return

    # ====== 檢查快取是否最新交易日的資料 ======
    all_max_dates = [df.index.max().date() for df in all_df.values() if len(df) > 0]
    today = max(all_max_dates) if all_max_dates else None
    if today:
        print(f"✅ 快取最新交易日為 {today}, 昨日訊號才有可能出現。")
    else:
        print("⚠️ 所有快取都沒有資料，昨日訊號一定為空。")

    # ================== 1.Random Forest 訓練 ==================
    print("\n[Step 1: Random Forest 訓練權重 (Random Forest Algo)]")
    t_rf = time.time()
    process_training_data_with_args = partial(process_training_data, all_df=all_df)

    # 使用多進程平行處理訓練資料
    with mp.get_context('spawn').Pool(processes=min(cpu_count(), 4)) as pool:
        results = list(tqdm(pool.imap(process_training_data_with_args, tw_list), total=len(tw_list), desc="組合訓練資料"))
    X_all, y_all = [], []
    for X, y in results:
        X_all.extend(X)
        y_all.extend(y)

    from sklearn.utils import resample
    Xy = list(zip(X_all,y_all))
    Xy_majority = [xy for xy in Xy if xy[1] == 0]
    Xy_minority = [xy for xy in Xy if xy[1] == 1]
    Xy_minority_upsampled = resample(Xy_minority, replace=True, n_samples=len(Xy_majority), random_state=42)
    Xy_balanced = Xy_majority + Xy_minority_upsampled
    X_all_balenced, y_all_balanced = zip(*Xy_balanced)

    scaler = StandardScaler()
    X_all_balenced = scaler.fit_transform(X_all_balenced)
    X_train, X_test, y_train, y_test = train_test_split(X_all_balenced, y_all_balanced, test_size=0.3, random_state=42, stratify=y_all_balanced)
    clf = RandomForestClassifier( # Random Forest 用這些 X，y來「學習」：什麼樣的特徵組合，容易在未來幾天內達到獲利目標
        n_estimators=100, # 森林中樹的數量，越多樹通常泛化能力越好(太少樹容易不穩定)，但訓練/預測會變慢 (100~500, GPT推薦300)
        random_state=42, # 隨機種子，讓每次結果一致。對模型本身表現沒有影響，只是方便重現
        max_depth = 8, # 每棵樹的最大深度(讓模型只學到較大方向的規律，減少記住雜訊) => 越大越容易過擬合，越小越容易欠擬合(4~12 或 None, GPT推薦8)
        min_samples_split = 10, # 一個節點要分割時，至少要有這麼多樣本。數字越大，樹越不容易長太深 => 可以防止樹對小樣本過度擬合(2~20)
        min_samples_leaf = 3, # 每個葉節點至少要有這麼多樣本。數字越大，樹越不容易長太細 => 防止樹對極端個案過擬合(1~10)
        max_features='log2', # 控制每次節點分裂時可用的特徵數量，增加隨機性。不是每次都用全部特徵，而是隨機選一部分特徵來決定最佳分裂點。減少樹之間的相關性，讓每棵樹學到不一樣的東西 => 提升整體泛化能力，降低過擬合('sqrt','log2',None)
        class_weight='balanced'
    )
    clf.fit(X_train, y_train) # 用 years 年的資料，訓練出一個能預測「未來幾天內會不會漲到目標」的模型
    print(f"訓練集勝率: {clf.score(X_train, y_train):.2%}")
    print(f"測試集勝率: {clf.score(X_test, y_test):.2%}")
    print(f"Random Forest 訓練耗時: {time.time() - t_rf:.2f} 秒")

    print("y_all 樣本分布 (0=敗, 1=勝):")
    print(pd.Series(y_all).value_counts())
    print("勝率比例: {:2%}".format(np.mean(y_all)))
    from sklearn.metrics import confusion_matrix, classification_report

    y_pred = clf.predict(X_test)
    print("\n混淆矩陣 (Confusion Matrix):")
    print(confusion_matrix(y_test, y_pred))
    print("\n分類報告 (Classification Report):")
    print(classification_report(y_test, y_pred, digits=4))

    importances = clf.feature_importances_
    print("\n各特徵對模型預測勝率的重要性: ")
    for name, imp in zip(feature_names, importances):
        print(f"{name:12s}: {imp:.4f}")

    plt.figure(figsize=(10,5))
    plt.bar(feature_names, importances)
    plt.xticks(rotation=45)
    plt.title("Feature Importances")
    plt.ylabel("Importance")
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, "feature_importances.png"))
    print(f"特徵重要性圖已存檔於 {os.path.join(save_dir, 'feature_importances.png')}")

    # ================== 2.多股票回測與分數分布 ==================
    print("\n[Step 2: 多股票回測與分數分布 (Multi Stock Analyze)]")
    t_rt = time.time()
    process_backtest_with_args = partial(process_backtest, all_df=all_df, clf=clf, scaler=scaler)

    # 使用多進程處理回測
    with mp.get_context('spawn').Pool(processes=min(cpu_count(), 4)) as pool:
        all_results = list(tqdm(pool.imap(process_backtest_with_args, tw_list), total=len(tw_list), desc="回測分數分布"))
    all_scores_info = [item for results in all_results for item in results]
    print(f"回測分數分布總耗時: {time.time() - t_rt:.2f} 秒")

    # ====== 統一計算 95 分位數 ======
    all_scores = np.array ([x[2] for x in all_scores_info])
    score_95 = np.percentile(all_scores, 95) if len(all_scores) > 0 else 0.5
    print(f"\n分數分布統計: 最小值：{all_scores.min():.2f}，最大值: {all_scores.max():.2f}，平均值:{all_scores.mean():.2f}，中位數:{np.median(all_scores):.2f}，95分位: {score_95:.2f}")

    # ====== 依據 95 分位數門檻，且同一股票進場後 10 天內不再進場 ======
    signal_recent = []
    recent_dates = sorted(set(x[1] for x in all_scores_info if isinstance(x[1], pd.Timestamp)))[-show_days:]
    for symbol in set(x[0] for x in all_scores_info):
        symbol_scores = [x for x in all_scores_info if x[0] == symbol and x[1] in recent_dates and x[2] >= score_95]
        symbol_scores.sort(key=lambda x: x[1]) # 按日期排序
        i = 0
        while i < len(symbol_scores):
            s = symbol_scores[i]
            features = s[-1]
            bullish_ma_n = features[feature_names.index('bullish_ma_n')]
            vol_up_n = features[feature_names.index('vol_up_n')]
            bullish_ma_and_rsi = features[feature_names.index('bullish_ma_and_rsi')]
            # 預留加權/分數門檻調整空間
            # s_score = s[2]
            # if bullish_ma_n == 1: s_score *= 1.1
            # if vol_up_n == 1: s_score *= 1.05
            # if bullish_ma_and_rsi == 1: s_score *= 1.1
            if bullish_ma_n == 1 or vol_up_n == 1 or bullish_ma_and_rsi == 1:
                signal_recent.append(s)
            i += 1
            next_allowed_date = s[1] + pd.Timedelta(days=lookahead_days)
            while i < len(symbol_scores) and symbol_scores[i][1] < next_allowed_date:
                i += 1

    print(f"\n近 {show_days} 天内分數 >= 95分位 ({score_95:.2f}) 且多頭排列持續且量能放大且多頭+RSI>50的進場訊號:")
    win_count = sum(s[3] for s in signal_recent)
    for symbol, date, score, win, profit_pct, max_profit_date, features in signal_recent:
        features_fmt = [f"{int(f)}" if isinstance(f, (int, np.integer)) or (isinstance(f, float) and f.is_integer()) else f"{f:.2f}" for f in features]
        if max_profit_date is not None:
            print(f"{symbol} {date:%Y-%m-%d} features: {features_fmt} 分數: {score:.2f}進場 {'勝'if win else '敗'}，{lookahead_days} 天內最高獲利: {profit_pct:.2f}% ({max_profit_date:%Y-%m-%d})")
        else:
            print(f"{symbol} {date:%Y-%m-%d} features: {features_fmt} 分數: {score:.2f}進場 {'勝'if win else '敗'}，{lookahead_days} 天內最高獲利: {profit_pct:.2f}%")
    total_count = len(signal_recent)
    print(f"\n近 {show_days} 天內總勝率: {win_count}/{total_count} = {win_count/total_count*100:.2f}%" if total_count > 0 else f"\n近 {show_days} 天內無進場訊號")

    # ====== 勝率分析 ======
    # 分數 >= 95分位 且 bullish_ma_n = 1
    highscore_bullish_win = []
    for symbol, date, score, win, profit_pct, max_profit_date, features in all_scores_info:
        if score >= score_95 and features[feature_names.index('bullish_ma_n')] == 1:
            highscore_bullish_win.append(win)
    print(f"分數>95分位且多頭排列持續性勝率: {np.mean(highscore_bullish_win):.2%} (樣本數: {len(highscore_bullish_win)})")
    
    # 量能持續放大
    bullish_and_vol_win = []
    for symbol, date, score, win, profit_pct, max_profit_date, features in all_scores_info:
        if score >= score_95 and features[feature_names.index('vol_up_n')] == 1:
            bullish_and_vol_win.append(win)
    print(f"分數>95分位且多頭排列且量能持續放大勝率: {np.mean(bullish_and_vol_win):.2%} (樣本數: {len(bullish_and_vol_win)})")

    # 剛形成多頭排列
    bullish_ma_just_win = []
    for symbol, date, score, win, profit_pct, max_profit_date, features in all_scores_info:
        idx_just = feature_names.index('bullish_ma_just')
        if features[idx_just] == 1:
            bullish_ma_just_win.append(win)
    print(f"剛形成多頭排列勝率: {np.mean(bullish_ma_just_win):.2%} (樣本數: {len(bullish_ma_just_win)})")

    signals_df = pd.DataFrame(
        all_scores_info,
        columns=['stock_id', 'date', 'score', 'win', 'profit_pct', 'max_profit_date', 'features']
    )
    signals_df['date'] = pd.to_datetime(signals_df['date']).dt.date
    # print("signals_df['date'] unique:", signals_df['date'].unique())

    if is_plot:
        # ================== 3. 單股畫圖與存圖 (只畫 show_days 內的訊號) ==================
        print("\n[Step 3: 單股畫圖與存圖 (Single Stock Analyze and plot)]")
        t_plot = time.time()

        # 先把每個股票的所有進場訊號日期都收集起來
        stock_signal_dates = defaultdict(list)
        for symbol, date, *_ in signal_recent:
            stock_signal_dates[symbol].append(date)

        # 所有有快取的股票都畫圖，紅圈圈指標有進場訊號的日期
        for stock_symbol in tqdm(all_df.keys(), desc="畫圖與存圖"):
            df = all_df[stock_symbol]
            if len(df) > show_days:
                df = df.iloc[-show_days:]
            save_path = os.path.join(save_dir, f"{stock_symbol}_signals.png")
            try:
                plot_signals(df, clf, scaler, score_95, save_path, show_days, vol_times)
            except Exception as e:
                print(f"存圖失敗：{e}，路徑: {save_path}")
        print(f"畫圖總耗時: {time.time() - t_plot:.2f} 秒")
    else:
        print(f"\n[BYPASS Step3: 畫圖]")

    # ================== 4. 進出場買賣模擬 ==================
    print("\n[Step 4: 進出場買賣模擬 (Entry and exit trading simulation)]")
    
    # 取得所有交易日序列
    all_dates = sorted(set(date for df in all_df.values() for date in df.index.date))
    if len(all_dates) >= 2:
        prev_trade_date = all_dates[-2]
        today = all_dates[-1]
    else:
        prev_trade_date = today = None

    # 產生今日有訊號的 all_signals_today
    if today:
        all_signals_today = signals_df[signals_df['date'] == today].values.tolist()
    else:
        all_signals_today = []

    print(f"\n🔍 昨日 ({prev_trade_date}) 有訊號，今日 ({today}) 用開盤價進場 (模擬持有部位)")
    today_signal_list = []
    if prev_trade_date and today:
        for stock_symbol, df in all_df.items():
            if prev_trade_date not in df.index.date or today not in df.index.date:
                continue
            prev_idx = df.index.get_loc(pd.to_datetime(prev_trade_date))
            f = make_features(df, prev_idx)
            if f is None:
                continue
            s = clf.predict_proba(scaler.transform([f]))[0][1]
            bullish_ma_n = f[feature_names.index('bullish_ma_n')]
            vol_up_n = f[feature_names.index('vol_up_n')]
            bullish_ma_and_rsi = f[feature_names.index('bullish_ma_and_rsi')]
            # 預留加權/分數門檻調整空間
            # s_score = s[2]
            # if bullish_ma_n == 1: s_score *= 1.1
            # if vol_up_n == 1: s_score *= 1.05
            # if bullish_ma_and_rsi == 1: s_score *= 1.1
            if s >= score_95 and (bullish_ma_n == 1 or vol_up_n == 1 or bullish_ma_and_rsi == 1):
                entry_date = df.index[df.index.get_loc(pd.to_datetime(today))]
                entry_price = df.loc[entry_date, 'open']
                today_signal_list.append((stock_symbol, entry_date, s, entry_price))
                stock_name = twstock.codes.get(str(stock_symbol), None)
                stock_name = stock_name.name if stock_name else ''
                print(f"模擬進場 {stock_symbol} {stock_name} 昨日 ({prev_trade_date}) 分數: {s:.3f}，今日({entry_date.date()}) 以開盤價 {entry_price} 進場 ✅！")
        if not today_signal_list:
            print("昨日所有股票都沒有進場訊號")
    else:
        print("無法取得最新交易日，無法產生進場訊號")

    # 取得最近 10 天的日期
    all_dates_dt = set()
    for df in all_df.values():
        all_dates_dt.update(df.index)
    recent_10_dates = sorted(all_dates_dt)[-10:]
    all_signals_10d = [x for x in all_scores_info if x[1] in recent_10_dates and x[2] >= score_95 and
                       (x[-1][feature_names.index('bullish_ma_n')] == 1 and 
                        x[-1][feature_names.index('vol_up_n')] == 1 and 
                        x[-1][feature_names.index('bullish_ma_and_rsi')] == 1)]

    # ====== 補上最後 lookahead_days 天的預測訊號（無勝負結果）======
    for stock_symbol, df in all_df.items():
        if len(df) < 61:
            continue
        for i in range(len(df) - lookahead_days, len(df)):
            date = df.index[i]
            if date not in recent_10_dates:
                continue
            f = make_features(df, i)
            if f is None:
                continue
            s = clf.predict_proba(scaler.transform([f]))[0][1]
            bullish_ma_n = f[feature_names.index('bullish_ma_n')]
            vol_up_n = f[feature_names.index('vol_up_n')]
            bullish_ma_and_rsi = f[feature_names.index('bullish_ma_and_rsi')]
            if s >= score_95 and (bullish_ma_n == 1 or vol_up_n == 1 or bullish_ma_and_rsi == 1):
                # 檢查是否已經在 all_signals_10d
                if not any((x[0] == stock_symbol and x[1] == date) for x in all_signals_10d):
                    all_signals_10d.append((stock_symbol, date, s, None, None, None, f))
    
    # 改成依股票號碼排序
    all_signals_10d = sorted(all_signals_10d, key=lambda x: int(x[0]))

    print(f"\n最近 10 天內所有分數 >= 95 分位({score_95:.2f}) 且多頭排列持續且量能放大且多頭+RSI>50的進場訊號 (含預測，無勝負結果)")
    if all_signals_10d:
        for symbol, date, score, win, profit_pct, max_profit_date, features in all_signals_10d:
            stock_name = twstock.codes.get(str(stock_symbol), None)
            stock_name = stock_name.name if stock_name else ''
            if win is None:
                print(f"{symbol} {stock_name} {date:%Y-%m-%d} 分數: {score:.3f} 預測訊號，無勝負結果")
            else:
                print(f"{symbol} {stock_name} {date:%Y-%m-%d} 分數: {score:.3f} 進場 {'勝' if win else '敗'}，最高獲利: {profit_pct:.2f}%")
    else:
        print("最近 10 天內沒有任何進場訊號")

    # 印 Predicted 區塊 (今日有訊號，預測 next 交易日進場)
    messages = []
    predicted_today = [x for x in all_signals_10d if pd.to_datetime(x[1]).date() == today]
    next_trade_date = next_trading_day(today)
    print(f"\n今日({today})有訊號，預測({next_trade_date.date()})進場的 Predicted")
    for symbol, date, score, win, profit_pct, max_profit_date, features in predicted_today:
        stock_name = twstock.codes.get(str(symbol), None)
        stock_name = stock_name.name if stock_name else ''
        print(f"{symbol} {stock_name} 今日 ({today}) 分數: {score:.3f} 下個交易日 ({next_trade_date.date()}) 預測進場")
        line = f"{symbol} {stock_name} 分數: {score:.3f}"
        messages.append(line)
            
    if messages:
        messages = (
            f"📢 產業別：{industry}\n"
            f"預測日期：{today:%Y-%m-%d}，預測進場日：{next_trade_date}\n\n"
            + "\n".join(messages)
        )
    else:
        messages = f"{industry} 昨日無符合條件的訊號"

    # response = send_line_message(messages)
    # print(f"\nline message: {response}")

    update_trade_excel(
        today_signal_list, all_df, global_excel_path, 
        sheet_name=user_choice, 
        profit_target=profit_target, 
        lookahead_days=lookahead_days, 
        all_signals_10d=all_signals_10d, 
        all_signals_today=all_signals_today
    )
    print(f"\n全部執行總耗時: {time.time() - start_time:.2f} 秒")

    return messages

if __name__ == '__main__':
    freeze_support()
    today_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    base_save_dir = r'C:\Users\Jerry-yc.Liu\Desktop\script\stock_signal_charts'

    cache_dir = r'C:\Users\Jerry-yc.Liu\Desktop\script\cache'
    os.makedirs(base_save_dir, exist_ok=True)
    os.makedirs(cache_dir, exist_ok=True)

    # 固定 Excel 檔案路徑 (只會有一份，持續追蹤部位)
    global_excel_path = os.path.join(base_save_dir, "auto_trade_sim.xlsx")

    final_messages = []
    # 支援 user_choice = 'all'
    if user_choice == 'all':
        for industry in industry_dict.keys():
            print(f"\n========= 處理 {industry} ===========")
            globals()['user_choice'] = industry
            save_dir = os.path.join(base_save_dir, f"{today_str}_{industry}")
            log_path = os.path.join(save_dir, f"stock_analysis_{today_str}_{industry}.log")
            os.makedirs(save_dir, exist_ok=True)
            old_stdout = sys.stdout
            sys.stdout = Logger(log_path)
            try:
                industry_message = main(save_dir, cache_dir, log_path, global_excel_path, industry)
                final_messages.append(industry_message)
            finally:
                sys.stdout = old_stdout
        # ✅ 所有產業跑完後，一次性發送
        final_message = "\n\n" + ("=" * 30) + "\n\n"
        final_message = final_message.join(final_messages)
        response = send_line_message(final_message)
        print(f"\nLine message sent: {response}")
    else:
        save_dir = os.path.join(base_save_dir, f"{today_str}_{user_choice}")
        log_path = os.path.join(save_dir, f"stock_analysis_{today_str}.log")
        os.makedirs(save_dir, exist_ok=True)
        message = main(save_dir, cache_dir, log_path, global_excel_path, user_choice)
        response = send_line_message(message)
        print(f"\nLine message sent: {response}")
    